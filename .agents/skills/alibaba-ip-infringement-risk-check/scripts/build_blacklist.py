"""
从 品牌层_知产高风险知识库.xlsx 提取核心字段并生成精简 CSV 黑名单。

用法：
    py build_blacklist.py <input.xlsx> <output.csv>

输出字段（9 列）：
    brand_name, brand_id, rights_owner, complaint_email,
    rights_type, rights_name, rights_id, category, risk_level
"""
from __future__ import annotations
import csv
import sys
import warnings
from pathlib import Path

import openpyxl  # type: ignore

warnings.filterwarnings("ignore")


# Excel 列索引（1-based） — 来自实际表头检查
COL = {
    "brand_name": 2,        # 品牌名称
    "brand_id": 3,          # 内部品牌编号
    "rights_owner": 4,      # 权利人名称
    "complaint_email": 5,   # 投诉邮箱
    "rights_type": 7,       # 权利类型
    "rights_name": 8,       # 权利名称
    "rights_id": 9,         # 权利编号
    "category": 10,         # 管控大类
    "enabled": 11,          # 品牌是否启用
    "risk_level": 15,       # 风险等级
    "status": 23,           # 实例状态
}

DATA_START_ROW = 3   # 前 2 行是表头


def main(xlsx_path: str, csv_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["数据"]

    rows_out: list[dict] = []
    seen: set[str] = set()
    skipped_disabled = 0
    skipped_dup = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        def get(c: int) -> str:
            return ws.cell(row=row_idx, column=c).value

        brand = (get(COL["brand_name"]) or "").strip()
        if not brand:
            continue

        enabled = (get(COL["enabled"]) or "").strip()
        # 只保留启用且审批通过的品牌
        if enabled and enabled != "是":
            skipped_disabled += 1
            continue

        # 去重 key：品牌名（不区分大小写）
        key = brand.lower()
        if key in seen:
            skipped_dup += 1
            continue
        seen.add(key)

        rows_out.append({
            "brand_name": brand,
            "brand_id": (get(COL["brand_id"]) or "").strip(),
            "rights_owner": (get(COL["rights_owner"]) or "").strip(),
            "complaint_email": (get(COL["complaint_email"]) or "").strip(),
            "rights_type": (get(COL["rights_type"]) or "").strip(),
            "rights_name": (get(COL["rights_name"]) or "").strip(),
            "rights_id": (get(COL["rights_id"]) or "").strip(),
            "category": (get(COL["category"]) or "").strip(),
            "risk_level": (get(COL["risk_level"]) or "").strip(),
        })

    # 按品牌名排序，便于查阅
    rows_out.sort(key=lambda r: r["brand_name"].lower())

    fieldnames = [
        "brand_name", "brand_id", "rights_owner", "complaint_email",
        "rights_type", "rights_name", "rights_id", "category", "risk_level",
    ]
    out_path = Path(csv_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows_out)

    print(f"[OK] {len(rows_out)} brands -> {csv_path}")
    print(f"     skipped_disabled={skipped_disabled}, skipped_duplicate={skipped_dup}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
